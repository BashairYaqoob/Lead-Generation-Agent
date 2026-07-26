"""Excel export module.

Writes a list of collected Lead objects to a formatted .xlsx file using
openpyxl: bold + frozen header row, auto-filter, auto-sized columns, blank
values left empty, and full Unicode preserved (native to openpyxl/xlsx).
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from models import Lead

_COLUMN_HEADERS = ["Business Name", "Email", "Phone Number", "Website", "Location"]

_MIN_COLUMN_WIDTH = 12
_MAX_COLUMN_WIDTH = 60


class ExcelExporter:
    """Exports a list of Lead objects to a formatted Excel (.xlsx) file."""

    def export(self, leads: list[Lead], filename: str) -> str:
        """Export leads to an Excel file with headers, filters, and sizing.

        Args:
            leads: The list of Lead objects to export (may be empty).
            filename: The desired output filename or path.

        Returns:
            The full path to the saved Excel file.

        Raises:
            OSError: If the file cannot be written.
        """
        full_path = self._resolve_output_path(filename)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Leads"

        self._write_headers(worksheet)
        self._write_rows(worksheet, leads)
        self._autosize_columns(worksheet)
        self._apply_header_formatting(worksheet, len(leads))

        os.makedirs(os.path.dirname(full_path), exist_ok=True)

        try:
            workbook.save(full_path)
        except OSError as exc:
            raise OSError(f"Failed to save Excel file to '{full_path}': {exc}") from exc

        return full_path

    @staticmethod
    def _resolve_output_path(filename: str) -> str:
        """Resolve the final output file path for the given filename.

        Args:
            filename: The desired filename, with or without a directory
                or extension.

        Returns:
            A path ending in the output directory and ".xlsx".
        """
        import config

        name = filename.strip()
        if not name.lower().endswith(".xlsx"):
            name = f"{name}.xlsx"

        directory = os.path.dirname(name)
        if directory:
            return name

        return os.path.join(config.OUTPUT_DIRECTORY, name)

    @staticmethod
    def _write_headers(worksheet: Worksheet) -> None:
        """Write the header row with bold styling.

        Args:
            worksheet: The worksheet to write headers into.
        """
        for column_index, header in enumerate(_COLUMN_HEADERS, start=1):
            cell = worksheet.cell(row=1, column=column_index, value=header)
            cell.font = Font(bold=True)

    @staticmethod
    def _write_rows(worksheet: Worksheet, leads: list[Lead]) -> None:
        """Write one row per lead below the header row, leaving blanks empty.

        Args:
            worksheet: The worksheet to write rows into.
            leads: The list of leads to write.
        """
        for row_index, lead in enumerate(leads, start=2):
            data = lead.to_dict()
            for column_index, header in enumerate(_COLUMN_HEADERS, start=1):
                value = data.get(header, "")
                # Write None instead of "" so empty cells are truly blank
                # in Excel rather than containing a zero-length string.
                worksheet.cell(row=row_index, column=column_index, value=value or None)

    @staticmethod
    def _autosize_columns(worksheet: Worksheet) -> None:
        """Adjust each column's width to roughly fit its longest value.

        Args:
            worksheet: The worksheet whose columns should be resized.
        """
        for column_index, header in enumerate(_COLUMN_HEADERS, start=1):
            column_letter = get_column_letter(column_index)
            max_length = len(header)

            for row in worksheet.iter_rows(
                min_col=column_index, max_col=column_index, min_row=2
            ):
                cell_value = row[0].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            width = min(max(max_length + 2, _MIN_COLUMN_WIDTH), _MAX_COLUMN_WIDTH)
            worksheet.column_dimensions[column_letter].width = width

    @staticmethod
    def _apply_header_formatting(worksheet: Worksheet, row_count: int) -> None:
        """Freeze the header row and enable auto-filter on the data range.

        Args:
            worksheet: The worksheet to format.
            row_count: The number of data rows written (excluding header).
        """
        worksheet.freeze_panes = "A2"

        last_column_letter = get_column_letter(len(_COLUMN_HEADERS))
        last_row = max(row_count + 1, 1)
        worksheet.auto_filter.ref = f"A1:{last_column_letter}{last_row}"